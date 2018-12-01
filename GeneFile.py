"""
GeneMark Query
Author: Matthew Lazeroff
"""

import requests
from bs4 import BeautifulSoup
import time
import os
import sys

# Genemark Domains
FILE_DOMAIN = 'http://exon.biology.gatech.edu/GeneMark/'
GM_DOMAIN = 'http://exon.gatech.edu/GeneMark/gm.cgi'
GM_HMM_DOMAIN = 'http://exon.gatech.edu/GeneMark/gmhmmp.cgi'
GMS_DOMAIN = 'http://exon.gatech.edu/GeneMark/genemarks.cgi'
HEURISTIC_DOMAIN = 'http://exon.gatech.edu/GeneMark/heuristic_gmhmmp.cgi'
GMS2_DOMAIN = 'http://exon.gatech.edu/GeneMark/genemarks2.cgi'
GLIMMER_DOMAIN = 'http://18.220.233.194/glimmer'


class Error(Exception):
    """
    Base Error Class
    """
    pass


class GeneFileError(Error):
    """
    Raised when no result from GeneMark is returned
    """

    def __init__(self, message):
        self.message = message


class GeneFile:
    """
    Class for querying GeneMark tools for a DNA sequence
    """

    def __init__(self, sequence_file):
        """
        Constructor
        Generates necessary parameters for post requests from DNA fasta file
        :param sequence_file:
        """
        # Load DNA Sequence into memory
        input_file_data = b''
        with open(sequence_file, 'rb') as input_file:
            current_byte = input_file.read(1)
            while (current_byte):
                input_file_data += current_byte
                current_byte = input_file.read(1)

        # get base file name
        self.name = str(os.path.basename(sequence_file).split('.')[0])

        # File creation for post requests
        self.file_info = {'file': (self.name, input_file_data, 'application/octet-stream')}

    def glimmer_query(self, out=''):
        """
        Queries Glimmer for DNA sequence and writes content to file
        :param out: optional output file name
        :return: name of file that was created
        """

        # parameters
        payload = [('sequence', self.file_info['file'][1]),
                   ('gencode', b'11'), ('topology', b'0'),
                   ('submit', b'Run GLIMMER v3.02')]

        # headers
        headers = {'User-Agent': 'GeneQuery'}

        # perform POST of file data
        file_post = requests.post(GLIMMER_DOMAIN, data=payload, headers=headers)
        file_post.raise_for_status()
        # check for job_key in response, if not raise error
        try:
            if 'job_key' not in file_post.text:
                raise GeneFileError("Glimmer POST #1: Invalid response")
        except GeneFileError:
            raise

        # get job_key from response
        job_key = file_post.text.split('=')
        payload = [(job_key[0], job_key[1])]

        # query server for output file
        # if output file is not ready, wait 2 seconds and requery
        try:
            time.sleep(2)
            return_post = requests.post(GLIMMER_DOMAIN, data=payload, headers=headers)
            return_post.raise_for_status()
            while 'DOCTYPE' not in return_post.text:
                time.sleep(2)
                return_post = requests.post(GLIMMER_DOMAIN, data=payload, headers=headers)
                return_post.raise_for_status()
        except requests.exceptions.HTTPError as e:
            raise GeneFileError(
                'Glimmer Server Error: Check for DNA for proper format or check server status')

        # write data to file
        if out != '':
            output = out
        else:
            output = self.name + '.glimmer'
        glimmer_output = open(output, 'wb')
        glimmer_output.write(return_post.content)

        # return file name
        return output

    def genemark_query(self, out=''):
        """
        Query to GeneMark
        :param out: optional output file name
        :return: name of file created
        """
        # Begin GeneMark Lookup -----------------------------------------------------
        gm_data = {'sequence': '', 'org': 'Escherichia_coli_K_12_substr__MG1655',
                   'submit': 'Start GeneMark', 'email': '', 'subject': 'GeneMark', 'windowsize': 96,
                   'stepsize': 12, 'threshold': 0.5, 'rbs': 'none', 'mode': 'native'}

        # GeneMark post - if unsuccessful, error thrown
        gm_post_request = requests.post(GM_DOMAIN, files=self.file_info, data=gm_data)
        gm_post_request.raise_for_status()
        soup = BeautifulSoup(gm_post_request.text, 'html.parser')

        # Get URL for gm output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFileError("GeneMark")
        except GeneFileError:
            raise

        # write gm response to file
        if out != '':
            output = out
        else:
            output = self.name + '.gm'
        gm_output = open(output, 'wb')
        getGMFile = requests.get(FILE_DOMAIN + file_location)
        getGMFile.raise_for_status()
        gm_output.write(getGMFile.content)

        return output
        # End GeneMark Lookup -------------------------------------------------------

    def genemarkhmm_query(self, out=''):
        """
        Query GeneMark Hmm
        :param out: optional output file name
        :return: name of file created
        """
        # Begin GeneMark Hmm Lookup -------------------------------------------------
        gm_hmm_data = {'sequence': '', 'org': 'Escherichia_coli_K_12_substr__MG1655',
                       'submit': 'Start GeneMark.hmm', 'format': 'LST',
                       'subject': 'GeneMark.hmm prokaryotic',
                       'email': ''}

        # GeneMark hmm post - if unsuccessful, error thrown
        hmm_post_request = requests.post(GM_HMM_DOMAIN, files=self.file_info, data=gm_hmm_data)
        hmm_post_request.raise_for_status()
        soup = BeautifulSoup(hmm_post_request.text, 'html.parser')

        # Get URL for hmm output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFileError("GeneMark Hmm")
        except GeneFileError:
            raise
        # write response to file
        if out != '':
            output = out
        else:
            output = self.name + '.gmhmm'
        gmhmm_output = open(output, 'wb')
        getHmmFile = requests.get(FILE_DOMAIN + file_location)
        getHmmFile.raise_for_status()
        gmhmm_output.write(getHmmFile.content)

        return output
        # End GeneMark Hmm Lookup --------------------------------------------------

    def genemarks_query(self, out=''):
        """
        Query GeneMarkS
        :param out: optional output file name
        :return: name of file created
        """
        # Begin GeneMarkS Lookup ---------------------------------------------------
        gms_data = {'sequence': '', 'submit': 'Start GeneMarkS', 'mode': 'prok', 'format': 'LST',
                    'subject': 'GeneMarkS', 'gcode': 11}

        # GeneMarkS post - if unsuccessful, error thrown
        gms_post_request = requests.post(GMS_DOMAIN, files=self.file_info, data=gms_data)
        gms_post_request.raise_for_status()
        soup = BeautifulSoup(gms_post_request.text, 'html.parser')

        # Get URL for hmm output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFileError("GeneMarkS")
        except GeneFileError:
            raise

        # write response to file
        if out != '':
            output = out
        else:
            output = self.name + '.gms'
        gms_output = open(output, 'wb')
        getGmsFile = requests.get(FILE_DOMAIN + file_location)
        getGmsFile.raise_for_status()
        gms_output.write(getGmsFile.content)

        return output
        # End GeneMarkS Lookup -----------------------------------------------------

    def genemark_heuristic_query(self, out=''):
        """
        Query GeneMark Heuristic
        :param out: optional output file name
        :return: name of file created
        """
        # Begin GeneMark Heuristic Lookup ------------------------------------------
        heuristic_data = {'sequence': '', 'submit': 'Start GeneMark.hmm', 'format': 'LST',
                          'email': '',
                          'subject': 'GeneMark.hmm', 'gcode': 11, 'strand': 'both',
                          'mod_type': 1999}

        # GeneMark Heuristic post - if unsuccessful, error thrown
        heuristic_post_request = requests.post(HEURISTIC_DOMAIN, files=self.file_info,
                                               data=heuristic_data)
        heuristic_post_request.raise_for_status()
        soup = BeautifulSoup(heuristic_post_request.text, 'html.parser')

        # Get URL for heuristic output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFileError("GeneMark Heuristic")
        except GeneFileError:
            raise

        # write response to file
        if out != '':
            output = out
        else:
            output = self.name + '.heuristic'
        heuristic_output = open(output, 'wb')
        getHeuristicFile = requests.get(FILE_DOMAIN + file_location)
        getHeuristicFile.raise_for_status()
        heuristic_output.write(getHeuristicFile.content)

        return output
        # End GeneMark Heuristic Lookup -------------------------------------------

    def genemarks2_query(self, out=''):
        """
        Query GeneMarkS2
        :param out: optional output file name
        :return: name of file created
        """
        # Begin GeneMarkS2 Lookup -------------------------------------------------
        gmms2_data = {'sequence': '', 'submit': 'GeneMarkS-2', 'mode': 'auto', 'format': 'lst',
                      'email': '', 'subject': 'GeneMarkS-2', 'gcode': 11}

        # GeneMarkS2 Post Request
        gmms2_post_request = requests.post(GMS2_DOMAIN, files=self.file_info, data=gmms2_data)
        gmms2_post_request.raise_for_status()
        soup = BeautifulSoup(gmms2_post_request.text, 'html.parser')

        # Get URL for GMS2 output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFileError("GeneMarkS2")
        except GeneFileError:
            raise

        # write response to file
        if out != '':
            output = out
        else:
            output = self.name + '.gms2'
        gms2_output = open(output, 'wb')
        getGMS2File = requests.get(FILE_DOMAIN + file_location)
        getGMS2File.raise_for_status()
        gms2_output.write(getGMS2File.content)

        return output
        # End GeneMarkS2 Lookup --------------------------------------------------

    def query_all(self, output=''):
        """
        Query: Glimmer, GeneMark, GeneMarkHmm, GeneMarkS, GeneMarkS2, and GeneMark Heuristic
        :output: optional output directory, otherwise outputs to GeneFile directory
        :return: list of files written to in the following order:
               Glimmer, GeneMark, GeneMark Hmm, GeneMarkS, GeneMarkS2, GeneMarkHeuristic
        """

        # list containing file output names
        files = []

        # Begin Queries
        print('Start Querying:')

        # Glimmer
        print('Glimmer...', end='', flush=True)
        files.append(self.glimmer_query(out=output + self.name + '.glimmer'))
        print('done', flush=True)

        # GeneMark
        print('GeneMark...', end='', flush=True)
        files.append(self.genemark_query(out=output + self.name + '.gm'))
        print('done', flush=True)

        # Hmm
        print('Hmm...', end='', flush=True)
        files.append(self.genemarkhmm_query(out=output + self.name + '.gmhmm'))
        print('done', flush=True)

        # GMS
        print('GMS...', end='', flush=True)
        files.append(self.genemarks_query(out=output + self.name + '.gms'))
        print('done', flush=True)

        # GMS2
        print('GMS2...', end='', flush=True)
        files.append(self.genemarks2_query(out=output + self.name + '.gms2'))
        print('done', flush=True)

        # Heuristic
        print('Heuristic...', end='', flush=True)
        files.append(self.genemark_heuristic_query(out=output + self.name + '.heuristic'))
        print('done', flush=True)

        if output != '':
            directory = output
        else:
            directory = os.getcwd()
        print('Files written to:', directory)

        # return file list
        return files


class GeneError(Error):
    def __init__(self, message):
        self.message = message


class Gene:
    """
    Class for representing a potential gene encoding
    """

    def __init__(self, start, stop, direction, identity=''):
        """
        Constructor
        :param start:  start codon (int)
        :param stop:   end codon (int)
        :param direction: +/-
        :param identity: optional identifier
        """
        self.start = int(start)
        self.stop = int(stop)
        self.identity = identity
        # check for proper direction
        try:
            if direction != '+' and direction != '-':
                raise GeneError("Direction must be + or -")
            else:
                self.direction = direction
        except GeneError:
            raise
        self.length = self.stop - self.start + 1

    def __eq__(self, other):
        """
        Checks if Genes can possibly represent the same Gene
        For + direction, True is stop codons are same
        For - direction, True if start codons are same
        :param other: Gene object
        :return: True / False
        """
        try:
            if isinstance(other, Gene):
                if self.direction == other.direction:
                    if self.direction == '+':
                        if self.stop == other.stop:
                            return True
                    else:  # - direction
                        if self.start == other.start:
                            return True
            else:
                raise GeneError("Gene Eq: Comparing object must be of Gene type")
        except GeneError:
            raise

        return False

    def __str__(self):
        """
        Overload of str() for printing
        :return: String containing gene information
        """
        return 'Direction: {} Start: {} Stop: {} Length: {}'.format(self.direction,
                                                                    self.start,
                                                                    self.stop,
                                                                    self.length)


class GeneParse:
    """
    Class for parsing GeneMark output files
    All methods are static
    """

    @staticmethod
    def parse_glimmer(glimmer_file, identity=''):
        """
        Parse Glimmer output file for Genes
        :param glimmer_file: Glimmer output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """

        file = open(glimmer_file, 'r')

        # skip first line
        curr_line = file.readline()

        # Get Gene data
        genes = []
        curr_line = file.readline()
        while 'html' not in curr_line:
            data = [x for x in curr_line.split(' ') if x != '']
            # get gene direction and create Gene
            if '+' in data[3]:
                genes.append(Gene(data[1], data[2], '+', identity=identity))
            else:
                genes.append(Gene(data[2], data[1], '-', identity=identity))

            # grab next line
            curr_line = file.readline()

        # return list of Genes
        return genes

    @staticmethod
    def parse_genemark(gm_file, identity=''):
        """
        Parse GeneMark output file for Genes
        :param gm_file: GeneMark output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        file = open(gm_file, 'r')

        # Skip until Regions of Interests
        curr_line = file.readline()
        while (' LEnd      REnd    Strand      Frame' not in curr_line):
            curr_line = file.readline()

        # read blank line
        curr_line = file.readlines(2)

        # Get gene data
        curr_line = file.readline()
        genes = []
        while curr_line != '\n':
            data = [x for x in curr_line.strip().split(' ') if x != '']
            # check gene orientation
            if data[2] == 'direct':
                direction = '+'
            else:
                direction = '-'
            # add to list
            genes.append(Gene(data[0], data[1], direction, identity=identity))
            curr_line = file.readline()

        return genes

    @staticmethod
    def parse_genemarkS(s_file, identity=''):
        """
        Parse GeneMarkS file for Gene data
        :param s_file: GeneMarkS file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        file = open(s_file, 'r')

        # begin parse
        current_line = file.readline()
        # Check if proper GeneMark file
        try:
            if 'GeneMark.hmm' not in current_line:
                raise GeneError("Hmm: Not a valid GeneMark Hmm file")
        except GeneError:
            raise

        # skip through until gene data
        while ('Gene    Strand    LeftEnd    RightEnd       Gene     Class' not in current_line):
            current_line = file.readline()
        current_line = file.readline()  # read blank line

        # arary for gene objs
        genes = []
        # get gene data
        current_line = file.readline()
        while current_line != '\n':
            data = [x for x in current_line.strip().split(' ') if x != '']
            genes.append(Gene(data[2], data[3], data[1], identity=identity))
            current_line = file.readline()

        return genes

    @staticmethod
    def parse_genemarkHmm(hmm_file, identity=''):
        """
        Parse GeneMark Hmm file for Gene data
        :param hmm_file:
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        file = open(hmm_file, 'r')

        # begin parse
        current_line = file.readline()
        # Check if proper GeneMark file
        try:
            if 'GeneMark.hmm' not in current_line:
                raise GeneError("Hmm: Not a valid GeneMark Hmm file")
        except GeneError:
            raise

        # skip through until gene data
        while ('Gene    Strand    LeftEnd    RightEnd       Gene     Class' not in current_line):
            current_line = file.readline()
        current_line = file.readline()  # read blank line

        # arary for gene objs
        genes = []
        # get gene data
        current_line = file.readline()
        while current_line != '\n':
            data = [x for x in current_line.strip().split(' ') if x != '']
            genes.append(Gene(data[2], data[3], data[1], identity=identity))
            current_line = file.readline()

        return genes

    @staticmethod
    def parse_genemarkHeuristic(heuristic_file, identity=''):
        """
        Parse GeneMark Heuristic file for Gene data
        :param heuristic_file: GeneMark Heuristic output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in order
        """
        file = open(heuristic_file, 'r')

        # begin parse
        current_line = file.readline()
        # Check if proper GeneMark file
        try:
            if 'GeneMark.hmm' not in current_line:
                raise GeneError("Hmm: Not a valid GeneMark Hmm file")
        except GeneError:
            raise

        # skip through until gene data
        while ('Gene    Strand    LeftEnd    RightEnd       Gene     Class' not in current_line):
            current_line = file.readline()
        current_line = file.readline()  # read blank line

        # arary for gene objs
        genes = []
        # get gene data
        current_line = file.readline()
        while current_line != '\n':
            data = [x for x in current_line.strip().split(' ') if x != '']
            genes.append(Gene(data[2], data[3], data[1], identity=identity))
            current_line = file.readline()

        return genes

    @staticmethod
    def parse_genemarkS2(s2_file, identity=''):
        """
        Parse GeneMark S2 file for Gene data
        :param s2_file: GeneMark S2 output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        file = open(s2_file, 'r')

        # skip until gene data
        curr_line = file.readline()
        while ('SequenceID' not in curr_line):
            curr_line = file.readline()

        # Get gene data
        genes = []
        curr_line = file.readline()
        while curr_line != '\n':
            data = [x for x in curr_line.strip().split(' ') if x != '']
            genes.append(Gene(data[2], data[3], data[1], identity=identity))
            curr_line = file.readline()

        return genes


if __name__ == '__main__':

    ##########################################################################
    # Imports and Functions

    import argparse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, colors

    # Functions
    def write_gene(gene, row, ws, indexes):
        """
        Writes the passed gene into its corresponding columns in the Excel worksheet
        :param gene: Gene object
        :param row: row to write to
        :param ws: openpyxl worksheet object
        :param indexes: dictionary of indexes, organized by gene.identity labels
        """
        left = indexes[gene.identity][0][0] + str(row)
        right = indexes[gene.identity][0][1] + str(row)
        indexes[gene.identity][1] += 1

        # write data
        curr_row = ws[left:right][0]
        curr_row[0].value = indexes[gene.identity][1]
        curr_row[1].value = gene.direction
        curr_row[1].alignment = Alignment(horizontal='center')
        curr_row[2].value = gene.start
        curr_row[3].value = gene.stop
        curr_row[4].value = gene.length

    def color_row(ws, row, color):
        """
        Colors the row according to passed color
        Changes text color to white
        :param ws: worksheet to alter
        :param row: row number
        :param color: openpyxl Fill profile
        """
        row = ws['A' + str(row):'AI' + str(row)][0]
        for cell in row:
            cell.fill = color
            cell.font = Font(color=colors.WHITE)

    def excel_write(output_directory, files, sequence):
        """
        Writes the content of the Gene files to a spreadsheet
        :param output_directory: directory to output to
        :param files: list of file names returned by GeneFile.query_all()
        :param sequence: GeneFile object
        """
        # output to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Gene Calls'

        indexes = dict()
        indexes['GLIMMER'] = [['A', 'E'], 0]
        indexes['GM'] = [['G', 'K'], 0]
        indexes['HMM'] = [['M', 'Q'], 0]
        indexes['GMS'] = [['S', 'W'], 0]
        indexes['GMS2'] = [['Y', 'AC'], 0]
        indexes['HEURISTIC'] = [['AE', 'AI'], 0]
        headers_names = ['No.', 'Direction', 'Start', 'Stop', 'Length']

        # Row colors according to # of same genes
        colors = dict()
        colors[6] = PatternFill(fgColor='215967', fill_type='solid')
        colors[5] = PatternFill(fgColor='31869b', fill_type='solid')
        colors[4] = PatternFill(fgColor='92cddc', fill_type='solid')
        colors[3] = PatternFill(fgColor='b7dee8', fill_type='solid')
        colors[2] = PatternFill(fgColor='daeef3', fill_type='solid')

        # Format Columns
        for x in indexes.items():
            value = x[1]
            first = ws[value[0][0] + str(1): value[0][1] + str(1)][0]
            # Gene Program Label
            for column in first:
                column.value = x[0]
                column.font = Font(bold=True)
                column.alignment = Alignment(horizontal='center')

            # Column Headers
            second = ws[value[0][0] + str(2): value[0][1] + str(2)][0]
            for index, column in enumerate(second):
                column.value = headers_names[index]
                column.alignment = Alignment(horizontal='center')

        # get genes
        glim = GeneParse.parse_glimmer(files[0], identity='GLIMMER')
        gm = GeneParse.parse_genemark(files[1], identity='GM')
        hmm = GeneParse.parse_genemarkHmm(files[2], identity='HMM')
        gms = GeneParse.parse_genemarkS(files[3], identity='GMS')
        gms2 = GeneParse.parse_genemarkS2(files[4], identity='GMS2')
        heuristic = GeneParse.parse_genemarkHeuristic(files[5], identity='HEURISTIC')
        total = sorted(glim + gm + hmm + gms + gms2 + heuristic, key=lambda x: x.start)

        # write to file
        row = 3
        count = 0
        genes_in_row = 1
        while count < len(total):
            curr_gene = total[count]
            # print same genes on the same line, different ones on different lines
            if count != 0:
                if curr_gene != total[count - 1]:
                    # color previous row according to same # of genes
                    if genes_in_row > 1:
                        color_row(ws, row, colors[genes_in_row])
                    # move to next row and reset genes count
                    row += 1
                    genes_in_row = 1
                else:
                    genes_in_row += 1

            write_gene(curr_gene, row, ws, indexes)
            count += 1

        # apply color for last row
            # color previous row according to same # of genes
            if genes_in_row > 1:
                color_row(ws, row, colors[genes_in_row])

        wb.save(output_directory + sequence.name + '.xlsx')

###############################################################################
    # BEGIN MAIN EXECUTION

    # parse arguments
    parser = argparse.ArgumentParser()
    parser.add_argument('sequence', help='DNA sequence file')
    parser.add_argument('output', help='Output directory')
    parser.add_argument('-rm', help='Remove all Glimmer/GeneMark output files', action='store_true')
    args = parser.parse_args()

    # create sequence
    sequence = GeneFile(args.sequence)

    # check if output directory exists, if not, create it
    if not os.path.isdir(args.output):
        os.mkdir(args.output)

    # query
    files = sequence.query_all(output=args.output)

    # write to Excel file
    excel_write(args.output, files, sequence)

    # If -rm flag given, remove all output files
    if args.rm:
        for file in files:
            os.remove(file)
